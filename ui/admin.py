"""
ui/admin.py — Clean Professional Admin Panel
"""

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QPushButton, QScrollArea, QSizePolicy, QDialog,
    QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QTabWidget, QMessageBox, QFileDialog, QAbstractItemView,
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QColor

from ui.base import (
    TRAIT_COLORS, AnimatedBar, font,
    get_lang, LANG_BUS,
)
from core.database import (
    get_all_users, get_user_summary, get_admin_stats,
    delete_user, reset_user_data, export_all_data, verify_admin,
)

ACCENT = '#2563eb'
RED    = '#ef4444'
GREEN  = '#10b981'
GOLD   = '#f59e0b'
PURPLE = '#7c3aed'


class AdminLoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Admin Access')
        self.setFixedSize(340, 200)
        self.setStyleSheet('background:#fff;')
        self._ok = False

        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 28, 28, 28)
        lay.setSpacing(12)

        title = QLabel('Admin Login')
        title.setFont(font(16, True))
        title.setStyleSheet('color:#0f172a;')
        lay.addWidget(title)

        sub = QLabel('Masukkan password untuk mengakses panel admin.')
        sub.setFont(font(9))
        sub.setStyleSheet('color:#94a3b8;')
        sub.setWordWrap(True)
        lay.addWidget(sub)

        self._pw = QLineEdit()
        self._pw.setEchoMode(QLineEdit.Password)
        self._pw.setPlaceholderText('Password...')
        self._pw.setFont(font(11))
        self._pw.setFixedHeight(40)
        self._pw.setStyleSheet(
            'QLineEdit{background:#fff;color:#0f172a;'
            'border:1.5px solid #e2e8f0;border-radius:8px;padding:0 12px;}'
            'QLineEdit:focus{border-color:#2563eb;}'
        )
        self._pw.returnPressed.connect(self._login)
        lay.addWidget(self._pw)

        self._err = QLabel('')
        self._err.setFont(font(9))
        self._err.setStyleSheet('color:#ef4444;')
        lay.addWidget(self._err)

        btn_row = QHBoxLayout()
        cancel = QPushButton('Batal')
        cancel.setFont(font(10))
        cancel.setFixedHeight(36)
        cancel.setCursor(Qt.PointingHandCursor)
        cancel.setStyleSheet(
            'QPushButton{background:#fff;color:#64748b;'
            'border:1px solid #e2e8f0;border-radius:6px;padding:0 16px;}'
            'QPushButton:hover{border-color:#94a3b8;}'
        )
        cancel.clicked.connect(self.reject)

        login = QPushButton('Masuk')
        login.setFont(font(10, True))
        login.setFixedHeight(36)
        login.setCursor(Qt.PointingHandCursor)
        login.setStyleSheet(
            f'QPushButton{{background:{ACCENT};color:#fff;border:none;border-radius:6px;padding:0 16px;}}'
            f'QPushButton:hover{{background:#1d4ed8;}}'
        )
        login.clicked.connect(self._login)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(login)
        lay.addLayout(btn_row)

    def _login(self):
        if verify_admin(self._pw.text()):
            self._ok = True
            self.accept()
        else:
            self._err.setText('Password salah.')
            self._pw.clear()

    def success(self): return self._ok


class StatCard(QFrame):
    def __init__(self, value: str, label: str, color: str, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f'QFrame{{background:#fff;border:1px solid #e2e8f0;border-radius:8px;'
            f'border-left:3px solid {color};}}'
        )
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(2)

        self._val_lbl = QLabel(value)
        self._val_lbl.setFont(font(24, True))
        self._val_lbl.setStyleSheet(f'color:{color};')
        lay.addWidget(self._val_lbl)

        lbl = QLabel(label)
        lbl.setFont(font(9))
        lbl.setStyleSheet('color:#94a3b8;')
        lay.addWidget(lbl)

    def update_value(self, v: str): self._val_lbl.setText(v)


class DashboardTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet('background:#f8fafc;')
        self._build()

    def _build(self):
        from ui.base import ScrollPage
        scroll = ScrollPage()
        inner  = scroll.inner_layout()
        inner.setSpacing(16)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0,0,0,0)
        lay.setSpacing(0)

        # Stats row
        stats_row = QHBoxLayout()
        stats_row.setSpacing(10)
        self._c_users    = StatCard('0', 'Total Pengguna',  ACCENT)
        self._c_sessions = StatCard('0', 'Total Sesi',      PURPLE)
        self._c_iq       = StatCard('0', 'Sesi IQ',         '#f97316')
        self._c_bf       = StatCard('0', 'Sesi Big Five',   GREEN)
        self._c_avg_iq   = StatCard('—', 'Rata-rata IQ',    GOLD)
        for c in [self._c_users, self._c_sessions, self._c_iq, self._c_bf, self._c_avg_iq]:
            stats_row.addWidget(c)
        stats_w = QWidget()
        stats_w.setLayout(stats_row)
        inner.addWidget(stats_w)

        # OCEAN
        ocean_frame = QFrame()
        ocean_frame.setStyleSheet('background:#fff;border:1px solid #e2e8f0;border-radius:8px;')
        olay = QVBoxLayout(ocean_frame)
        olay.setContentsMargins(20, 16, 20, 16)
        olay.setSpacing(10)

        ot = QLabel('Rata-rata Skor OCEAN')
        ot.setFont(font(11, True))
        ot.setStyleSheet('color:#0f172a;')
        olay.addWidget(ot)

        div = QFrame()
        div.setFixedHeight(1)
        div.setStyleSheet('background:#f1f5f9;border:none;')
        olay.addWidget(div)

        self._ocean_bars = {}
        for t in 'OCEAN':
            col = TRAIT_COLORS[t]
            row = QHBoxLayout()
            lbl = QLabel(t)
            lbl.setFont(font(9, True))
            lbl.setStyleSheet(f'color:{col};')
            lbl.setFixedWidth(16)
            bar = AnimatedBar(0, col, 8)
            val = QLabel('0')
            val.setFont(font(9, True))
            val.setStyleSheet(f'color:{col};')
            val.setFixedWidth(32)
            bar._val = val
            self._ocean_bars[t] = bar
            row.addWidget(lbl)
            row.addWidget(bar, 1)
            row.addWidget(val)
            olay.addLayout(row)
        inner.addWidget(ocean_frame)

        # IQ Distribution
        self._dist_frame = QFrame()
        self._dist_frame.setStyleSheet('background:#fff;border:1px solid #e2e8f0;border-radius:8px;')
        dlay = QVBoxLayout(self._dist_frame)
        dlay.setContentsMargins(20, 16, 20, 16)
        dlay.setSpacing(8)
        dt = QLabel('Distribusi Kategori IQ')
        dt.setFont(font(11, True))
        dt.setStyleSheet('color:#0f172a;')
        dlay.addWidget(dt)
        ddiv = QFrame()
        ddiv.setFixedHeight(1)
        ddiv.setStyleSheet('background:#f1f5f9;border:none;')
        dlay.addWidget(ddiv)
        self._dist_inner = dlay
        inner.addWidget(self._dist_frame)
        inner.addStretch()
        lay.addWidget(scroll, 1)

    def refresh(self):
        stats = get_admin_stats()
        self._c_users.update_value(str(stats['total_users']))
        self._c_sessions.update_value(str(stats['total_sessions']))
        self._c_iq.update_value(str(stats['iq_sessions']))
        self._c_bf.update_value(str(stats['bf_sessions']))
        self._c_avg_iq.update_value(str(stats['avg_iq']) if stats['avg_iq'] else '—')

        for t in 'OCEAN':
            val = stats['ocean_avgs'].get(t, 0)
            self._ocean_bars[t].set_value(val)
            self._ocean_bars[t]._val.setText(f'{val:.0f}')

        # Clear dist rows (keep title + divider = first 2)
        while self._dist_inner.count() > 2:
            item = self._dist_inner.takeAt(2)
            if item.widget(): item.widget().deleteLater()
            elif item.layout():
                while item.layout().count():
                    sub = item.layout().takeAt(0)
                    if sub.widget(): sub.widget().deleteLater()

        total_iq = sum(d['cnt'] for d in stats['iq_dist'])
        for d in stats['iq_dist']:
            pct = (d['cnt'] / total_iq * 100) if total_iq > 0 else 0
            row = QHBoxLayout()
            lbl = QLabel(d['level'])
            lbl.setFont(font(9))
            lbl.setStyleSheet('color:#374151;')
            lbl.setFixedWidth(130)
            bar = AnimatedBar(pct, ACCENT, 8)
            cnt = QLabel(f'{d["cnt"]} ({pct:.0f}%)')
            cnt.setFont(font(9))
            cnt.setStyleSheet('color:#94a3b8;')
            cnt.setFixedWidth(80)
            row.addWidget(lbl)
            row.addWidget(bar, 1)
            row.addWidget(cnt)
            self._dist_inner.addLayout(row)


class UserListTab(QWidget):
    open_user = pyqtSignal(int, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet('background:#f8fafc;')
        self._all_users = []
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        top = QHBoxLayout()
        self._search = QLineEdit()
        self._search.setPlaceholderText('Cari nama pengguna...')
        self._search.setFont(font(10))
        self._search.setFixedHeight(36)
        self._search.setStyleSheet(
            'QLineEdit{background:#fff;color:#0f172a;'
            'border:1px solid #e2e8f0;border-radius:6px;padding:0 12px;}'
            'QLineEdit:focus{border-color:#2563eb;}'
        )
        self._search.textChanged.connect(self._filter)
        top.addWidget(self._search, 1)

        refresh_btn = QPushButton('Refresh')
        refresh_btn.setFont(font(10))
        refresh_btn.setFixedHeight(36)
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(
            f'QPushButton{{background:#fff;color:{ACCENT};'
            f'border:1px solid #e2e8f0;border-radius:6px;padding:0 14px;}}'
            f'QPushButton:hover{{border-color:{ACCENT};}}'
        )
        refresh_btn.clicked.connect(self.refresh)
        top.addWidget(refresh_btn)
        lay.addLayout(top)

        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels(['Nama', 'Terdaftar', 'Sesi', 'Terakhir', '', ''])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for i in [1,2,3]: self._table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self._table.setColumnWidth(4, 80)
        self._table.setColumnWidth(5, 80)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setStyleSheet('''
            QTableWidget{background:#fff;border:1px solid #e2e8f0;
                gridline-color:#f1f5f9;font-size:10px;border-radius:8px;}
            QHeaderView::section{background:#f8fafc;color:#374151;
                font-weight:600;padding:8px;border:none;
                border-bottom:1px solid #e2e8f0;}
            QTableWidget::item:alternate{background:#f8fafc;}
            QTableWidget::item:selected{background:#eff6ff;color:#0f172a;}
        ''')
        lay.addWidget(self._table)

        bot = QHBoxLayout()
        self._total_lbl = QLabel('')
        self._total_lbl.setFont(font(9))
        self._total_lbl.setStyleSheet('color:#94a3b8;')
        bot.addWidget(self._total_lbl)
        bot.addStretch()
        export_btn = QPushButton('Export Excel')
        export_btn.setFont(font(10, True))
        export_btn.setFixedHeight(34)
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.setStyleSheet(
            f'QPushButton{{background:{GREEN};color:#fff;border:none;border-radius:6px;padding:0 16px;}}'
            f'QPushButton:hover{{background:#059669;}}'
        )
        export_btn.clicked.connect(self._export_excel)
        bot.addWidget(export_btn)
        lay.addLayout(bot)

    def refresh(self):
        self._all_users = get_all_users()
        self._populate(self._all_users)

    def _filter(self, text):
        self._populate([u for u in self._all_users if text.lower() in u['name'].lower()])

    def _populate(self, users):
        self._table.setRowCount(len(users))
        for i, u in enumerate(users):
            self._table.setItem(i, 0, QTableWidgetItem(u['name']))
            self._table.setItem(i, 1, QTableWidgetItem(str(u.get('created_at',''))[:10]))
            self._table.setItem(i, 2, QTableWidgetItem(str(u.get('total_sessions', 0))))
            self._table.setItem(i, 3, QTableWidgetItem(str(u.get('last_session','—'))[:16]))

            det = QPushButton('Detail')
            det.setFont(font(9))
            det.setCursor(Qt.PointingHandCursor)
            det.setStyleSheet(
                f'QPushButton{{background:{ACCENT}15;color:{ACCENT};'
                f'border:1px solid {ACCENT}33;border-radius:4px;padding:2px 8px;}}'
                f'QPushButton:hover{{background:{ACCENT}30;}}'
            )
            det.clicked.connect(lambda _, uid=u['id'], nm=u['name']: self.open_user.emit(uid, nm))
            self._table.setCellWidget(i, 4, det)

            dl = QPushButton('Hapus')
            dl.setFont(font(9))
            dl.setCursor(Qt.PointingHandCursor)
            dl.setStyleSheet(
                f'QPushButton{{background:{RED}15;color:{RED};'
                f'border:1px solid {RED}33;border-radius:4px;padding:2px 8px;}}'
                f'QPushButton:hover{{background:{RED}30;}}'
            )
            dl.clicked.connect(lambda _, uid=u['id'], nm=u['name']: self._confirm_delete(uid, nm))
            self._table.setCellWidget(i, 5, dl)
            self._table.setRowHeight(i, 38)

        self._total_lbl.setText(f'{len(users)} pengguna')

    def _confirm_delete(self, user_id, name):
        msg = QMessageBox(self)
        msg.setWindowTitle('Hapus Pengguna')
        msg.setIcon(QMessageBox.Warning)
        msg.setText(f'Hapus <b>{name}</b>? Semua data akan dihapus permanen.')
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        if msg.exec_() == QMessageBox.Yes:
            delete_user(user_id)
            self.refresh()

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, 'Export', 'assessment_data.xlsx', 'Excel (*.xlsx)')
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            data = export_all_data()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Assessment Data'
            headers = ['Nama','Terdaftar','Tipe Tes','Bahasa','Tanggal','Durasi','Dimensi','Mentah','Normalized','Persentil','Level']
            hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            hfill = PatternFill('solid', fgColor='0f172a')
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')
            for ri, d in enumerate(data, 2):
                vals = [d['user_name'], d['user_created'][:10] if d['user_created'] else '',
                        d['test_type'], d['lang'], d['taken_at'][:16] if d['taken_at'] else '',
                        d['duration_s'] or 0, d['dimension'], d['raw_score'],
                        round(d['normalized'],1), round(d['percentile'],1), d['level'] or '']
                for ci, v in enumerate(vals, 1):
                    ws.cell(ri, ci, v)
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(len(str(c.value or '')) for c in col)+4, 28)
            ws.freeze_panes = 'A2'
            wb.save(path)
            QMessageBox.information(self, 'Export', f'Berhasil disimpan:\n{path}')
        except ImportError:
            QMessageBox.warning(self, 'Error', 'pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, 'Error', str(e))


class UserDetailTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet('background:#f8fafc;')
        self._user_id = None
        self._user_name = ''
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        self._header = QLabel('Pilih user dari tab Pengguna.')
        self._header.setFont(font(11))
        self._header.setStyleSheet('color:#94a3b8;')
        lay.addWidget(self._header)

        self._table = QTableWidget()
        self._table.setColumnCount(5)
        self._table.setHorizontalHeaderLabels(['Tipe', 'Tanggal', 'Bahasa', 'Durasi', 'Skor Utama'])
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        for i in range(4): self._table.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.setStyleSheet('''
            QTableWidget{background:#fff;border:1px solid #e2e8f0;
                gridline-color:#f1f5f9;font-size:10px;border-radius:8px;}
            QHeaderView::section{background:#f8fafc;color:#374151;
                font-weight:600;padding:8px;border:none;border-bottom:1px solid #e2e8f0;}
            QTableWidget::item:alternate{background:#f8fafc;}
        ''')
        lay.addWidget(self._table)

        self._reset_btn = QPushButton('Reset Semua Data User Ini')
        self._reset_btn.setFont(font(10))
        self._reset_btn.setFixedHeight(34)
        self._reset_btn.setCursor(Qt.PointingHandCursor)
        self._reset_btn.setStyleSheet(
            f'QPushButton{{background:#fff;color:{RED};'
            f'border:1px solid {RED}44;border-radius:6px;padding:0 16px;}}'
            f'QPushButton:hover{{background:{RED}10;}}'
        )
        self._reset_btn.setVisible(False)
        self._reset_btn.clicked.connect(self._reset_user)
        lay.addWidget(self._reset_btn, 0, Qt.AlignLeft)

    def load_user(self, user_id, user_name):
        self._user_id = user_id
        self._user_name = user_name
        self._header.setText(f'Sesi — {user_name}')
        self._header.setStyleSheet('color:#0f172a;font-weight:bold;font-size:12px;')
        self._reset_btn.setVisible(True)

        sessions = get_user_summary(user_id)
        self._table.setRowCount(len(sessions))
        for i, s in enumerate(sessions):
            tt = s.get('test_type','')
            col = ACCENT if tt=='iq' else PURPLE
            badge = 'IQ' if tt=='iq' else 'BF'
            t0 = QTableWidgetItem(badge)
            t0.setForeground(QColor(col))
            t0.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(i, 0, t0)
            self._table.setItem(i, 1, QTableWidgetItem(str(s.get('taken_at',''))[:16]))
            self._table.setItem(i, 2, QTableWidgetItem((s.get('lang') or 'id').upper()))
            dur = s.get('duration_s',0) or 0
            self._table.setItem(i, 3, QTableWidgetItem(f'{dur//60}m {dur%60}s'))
            scores_str = s.get('scores_str','') or ''
            parts = [p for p in scores_str.split('|') if ':' in p and p.split(':')[0] in ('IQ','O','C','E','A','N')]
            self._table.setItem(i, 4, QTableWidgetItem('  '.join(parts[:6])))
            self._table.setRowHeight(i, 36)

    def _reset_user(self):
        if not self._user_id: return
        msg = QMessageBox(self)
        msg.setWindowTitle('Reset Data')
        msg.setIcon(QMessageBox.Warning)
        msg.setText(f'Reset semua sesi <b>{self._user_name}</b>? Akun tetap ada.')
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        if msg.exec_() == QMessageBox.Yes:
            reset_user_data(self._user_id)
            self.load_user(self._user_id, self._user_name)


class AdminPanel(QWidget):
    closed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Admin Panel — Assessment v5.0')
        self.setMinimumSize(1000, 660)
        self.setStyleSheet('background:#f8fafc;')
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Topbar
        topbar = QFrame()
        topbar.setFixedHeight(52)
        topbar.setStyleSheet('background:#0f172a;border-bottom:1px solid #1e293b;')
        tlay = QHBoxLayout(topbar)
        tlay.setContentsMargins(20, 0, 20, 0)

        logo = QFrame()
        logo.setFixedSize(32, 32)
        logo.setStyleSheet(f'background:{ACCENT};border-radius:7px;')
        ll = QVBoxLayout(logo)
        ll.setContentsMargins(0,0,0,0)
        li = QLabel('A')
        li.setFont(font(11, True))
        li.setStyleSheet('color:#fff;background:transparent;')
        li.setAlignment(Qt.AlignCenter)
        ll.addWidget(li)
        tlay.addWidget(logo)
        tlay.addSpacing(10)

        title = QLabel('Admin Panel')
        title.setFont(font(13, True))
        title.setStyleSheet('color:#f1f5f9;')
        tlay.addWidget(title)
        tlay.addStretch()

        close_btn = QPushButton('Tutup')
        close_btn.setFont(font(10))
        close_btn.setFixedHeight(32)
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.setStyleSheet(
            f'QPushButton{{background:transparent;color:{RED};border:1px solid {RED}44;border-radius:6px;padding:0 14px;}}'
            f'QPushButton:hover{{background:{RED}15;}}'
        )
        close_btn.clicked.connect(self._close)
        tlay.addWidget(close_btn)
        root.addWidget(topbar)

        self._tabs = QTabWidget()
        self._tabs.setStyleSheet('''
            QTabWidget::pane{border:none;background:#f8fafc;}
            QTabBar::tab{background:#fff;color:#94a3b8;padding:10px 20px;
                border:none;border-bottom:2px solid transparent;font-size:11px;font-weight:600;}
            QTabBar::tab:selected{color:#2563eb;border-bottom-color:#2563eb;}
            QTabBar::tab:hover{color:#374151;}
        ''')

        self._dash      = DashboardTab()
        self._user_list = UserListTab()
        self._user_det  = UserDetailTab()

        self._tabs.addTab(self._dash,      'Dashboard')
        self._tabs.addTab(self._user_list, 'Pengguna')
        self._tabs.addTab(self._user_det,  'Detail User')
        self._user_list.open_user.connect(self._open_detail)
        root.addWidget(self._tabs, 1)

    def _open_detail(self, uid, name):
        self._user_det.load_user(uid, name)
        self._tabs.setCurrentIndex(2)

    def showEvent(self, e):
        super().showEvent(e)
        self._dash.refresh()
        self._user_list.refresh()

    def _close(self):
        self.closed.emit()
        self.close()